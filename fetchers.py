"""
fetchers.py
-----------
Funciones que traen datos de cada fuente y los devuelven SIEMPRE en el mismo
formato "largo" (tidy): un DataFrame con columnas
    fecha (datetime) | valor (float)

Cada fuente tiene su propia función; el orquestador (run.py) elige cuál usar
según el campo "fuente" de cada indicador en el config.

Fuentes soportadas:
  - datos_gob      -> apis.datos.gob.ar/series (EMAE, IPC, monetarias, fiscal, empleo...)
  - argentinadatos -> api.argentinadatos.com   (inflación, riesgo país)
  - dolar          -> api.argentinadatos.com/v1/cotizaciones/dolares (histórico por casa)
  - bcra           -> api.bcra.gob.ar/estadisticas (reservas diarias, etc.)
  - rem_bcra       -> Excel histórico único del REM (bcra.gob.ar), expectativas de mercado
"""
from __future__ import annotations
import datetime
import io
import json
import re
import time
from pathlib import Path
import requests
import pandas as pd
import xlrd
import openpyxl
import pdfplumber

TIMEOUT = 30
# Sin tildes/caracteres no-ASCII a propósito: un User-Agent con "é" (mal codificado en la
# cabecera HTTP) causaba un 502 consistente y reproducible en api.bcra.gob.ar -- probado en
# vivo, era la causa real de que "Reservas internacionales (BCRA)" no se actualizara.
HEADERS = {"User-Agent": "coyuntura-tracker/1.0 (uso academico)"}


def _get(url: str, params: dict | None = None, reintentos: int = 3) -> requests.Response:
    """GET con reintentos simples y backoff."""
    ultimo_error = None
    for intento in range(reintentos):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=TIMEOUT)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            ultimo_error = e
            time.sleep(2 * (intento + 1))
    raise RuntimeError(f"Falló la consulta a {url}: {ultimo_error}")


# ---------------------------------------------------------------------------
# 1) apis.datos.gob.ar/series  -> el backbone: +30.000 series oficiales
# ---------------------------------------------------------------------------
LIMITE_PAGINA_DATOS_GOB = 5000  # tope máximo de filas por página que acepta la API (400 si se pide más)


def fetch_datos_gob(serie_id: str, start_date: str | None = None) -> pd.DataFrame:
    """
    Trae una serie de la API de Series de Tiempo de la Nación, paginando con
    el parámetro 'start' (offset de filas) hasta cubrir el 'count' total que
    informa la API. Necesario para series diarias largas (BADLAR, etc.) que
    ya superan las 5000 observaciones: sin paginar, quedan cortadas para
    siempre en la fila 5000 y el indicador se congela en silencio.

    serie_id : id de la serie (ej '143.3_NO_PR_2004_A_21'). Buscalos con buscar_series.py
    """
    url = "https://apis.datos.gob.ar/series/api/series/"
    data = []
    offset = 0
    total = None
    while total is None or offset < total:
        params = {"ids": serie_id, "format": "json", "limit": LIMITE_PAGINA_DATOS_GOB, "start": offset}
        if start_date:
            params["start_date"] = start_date
        payload = _get(url, params=params).json()
        pagina = payload.get("data", [])
        if not pagina:
            break
        data.extend(pagina)
        total = payload.get("count", len(data))
        offset += len(pagina)
        if len(pagina) < LIMITE_PAGINA_DATOS_GOB:
            break  # última página (vino incompleta)

    if not data:
        return pd.DataFrame(columns=["fecha", "valor"])

    df = pd.DataFrame(data, columns=["fecha", "valor"])
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    return df.dropna().drop_duplicates(subset="fecha").sort_values("fecha").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2) ArgentinaDatos -> inflación, riesgo país (endpoints tipo {fecha, valor})
# ---------------------------------------------------------------------------
def fetch_argentinadatos(endpoint: str, start_date: str | None = None) -> pd.DataFrame:
    """
    endpoint : ruta relativa, ej 'finanzas/indices/riesgo-pais'
               o 'finanzas/indices/inflacion'
    """
    url = f"https://api.argentinadatos.com/v1/{endpoint}"
    data = _get(url).json()
    if not data:
        return pd.DataFrame(columns=["fecha", "valor"])

    df = pd.DataFrame(data)
    # estos endpoints devuelven {'fecha': ..., 'valor': ...}
    df = df.rename(columns={"fecha": "fecha", "valor": "valor"})
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    df = df.dropna().sort_values("fecha")
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df[["fecha", "valor"]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 3) Dólar histórico por casa (ArgentinaDatos)
# ---------------------------------------------------------------------------
def fetch_dolar(casa: str, start_date: str | None = None) -> pd.DataFrame:
    """
    casa : oficial | blue | bolsa (MEP) | contadoconliqui (CCL) | mayorista | tarjeta | cripto
    Devuelve el valor de VENTA por fecha.
    """
    url = f"https://api.argentinadatos.com/v1/cotizaciones/dolares/{casa}"
    data = _get(url).json()
    if not data:
        return pd.DataFrame(columns=["fecha", "valor"])

    df = pd.DataFrame(data)
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["valor"] = pd.to_numeric(df["venta"], errors="coerce")
    df = df.dropna(subset=["valor"]).sort_values("fecha")
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df[["fecha", "valor"]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 4) BCRA -> reservas internacionales diarias, otros datos diarios
# ---------------------------------------------------------------------------
def _parsear_respuesta_bcra(data: dict) -> list[dict]:
    """Extrae filas {fecha, valor} de una respuesta de la API de estadísticas del BCRA."""
    filas = []
    results = data.get("results")
    if not results:
        return filas
    # v4.0: [{"idVariable":1,"detalle":[{fecha,valor},...]}].
    if isinstance(results, list) and len(results) > 0 and results[0].get("detalle"):
        items = [d for item in results for d in item.get("detalle", [])]
    else:
        items = results
    for item in items:
        try:
            fecha_str = item.get("fecha")
            valor = item.get("valor")
            if fecha_str and valor is not None:
                filas.append({"fecha": fecha_str, "valor": float(valor)})
        except (ValueError, KeyError, TypeError):
            pass
    return filas


def fetch_bcra(id_variable: int, start_date: str | None = None) -> pd.DataFrame:
    """
    Trae datos de la API del BCRA (Estadísticas Monetarias), sólo v4.0.

    id_variable: número de variable (ej: 1 para reservas internacionales diarias)
    Devuelve DataFrame con fecha (diaria) y valor.

    v3.0 y v2.0 quedaron confirmadas MUERTAS (410 Gone y 404 Not Found
    respectivamente, no "inestables" -- probado en vivo, nunca van a responder)
    y se sacaron del todo: antes cada corrida fallida desperdiciaba 12 de 18
    intentos pegándole a esos dos endpoints muertos en vez de reintentar contra
    la única versión viva. Resiliente a la inestabilidad conocida de v4.0 (502
    intermitentes, confirmado que a veces responde y a veces no incluso con la
    misma request repetida en minutos): prueba ventanas de fecha cada vez más
    cortas (2 años · 6 meses · 1 mes) y reintenta cada una varias veces con
    backoff. Si todo falla, loguea el error y devuelve un DataFrame vacío SIN
    tocar el histórico ya guardado (storage.py sólo agrega, nunca borra).
    """
    end_date = pd.Timestamp.today()
    ventanas_dias = [730, 180, 30]  # 2 años, 6 meses, 1 mes
    reintentos_por_ventana = 4
    filas = []
    ultimo_error = None

    for dias in ventanas_dias:
        desde = (end_date - pd.Timedelta(days=dias)).strftime("%Y-%m-%d")
        hasta = end_date.strftime("%Y-%m-%d")
        url = f"https://api.bcra.gob.ar/estadisticas/v4.0/monetarias/{id_variable}"
        params = {"desde": desde, "hasta": hasta, "limit": 1000}

        for intento in range(reintentos_por_ventana):
            try:
                try:
                    r = requests.get(url, params=params, headers=HEADERS, timeout=TIMEOUT, verify=True)
                    r.raise_for_status()
                except (requests.exceptions.SSLError, requests.exceptions.ConnectionError):
                    r = requests.get(url, params=params, headers=HEADERS, timeout=TIMEOUT, verify=False)
                    r.raise_for_status()
                filas = _parsear_respuesta_bcra(r.json())
                if filas:
                    break
            except (requests.RequestException, ValueError, KeyError, TypeError) as e:
                ultimo_error = f"v4.0 desde={desde}: {e}"
                time.sleep(2 * (intento + 1))  # backoff antes del próximo reintento
        if filas:
            break

    if not filas:
        print(f"  [ADVERTENCIA] fetch_bcra id_variable={id_variable}: sin datos tras probar "
              f"{len(ventanas_dias)} ventanas x {reintentos_por_ventana} reintentos (sólo v4.0, "
              f"v3.0/v2.0 confirmadas muertas). Último error: {ultimo_error}")
        return pd.DataFrame(columns=["fecha", "valor"])

    df = pd.DataFrame(filas)
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    df = df.dropna().sort_values("fecha")
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df[["fecha", "valor"]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 5) REM (BCRA) -> Relevamiento de Expectativas de Mercado
# ---------------------------------------------------------------------------
REM_XLSX_URL = ("https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/"
                "informes/historico-relevamiento-expectativas-mercado.xlsx")
REM_FRESCURA_DIAS = 7  # el BCRA publica una encuesta nueva por mes: con 1 vez por semana alcanza de sobra
_CACHE_DIR = Path(__file__).resolve().parent / "data"
_REM_CACHE_CSV = _CACHE_DIR / "_cache_rem.csv"
_REM_CACHE_META = _CACHE_DIR / "_cache_rem_meta.json"


def _descargar_y_transformar_rem() -> pd.DataFrame:
    """
    Descarga el Excel histórico ÚNICO del REM (misma URL fija, ~1,5 MB) y
    devuelve TODAS las variables/referencias de la hoja "Base de Datos
    Completa" ya reducidas a "expectativa a un mes vista": para cada
    encuesta, sólo el pronóstico cuyo Período es el mes INMEDIATAMENTE
    siguiente al de la encuesta. Se transforman todas las variables de una
    sola pasada (no sólo la pedida) para que la caché sirva de una sola
    descarga a cualquier indicador REM que se agregue más adelante.
    """
    r = _get(REM_XLSX_URL)
    with io.BytesIO(r.content) as buf:
        df = pd.read_excel(buf, sheet_name="Base de Datos Completa", skiprows=1, engine="openpyxl")
    df = df.iloc[:, :5]
    df.columns = ["encuesta", "variable", "referencia", "periodo", "mediana"]
    df["encuesta"] = pd.to_datetime(df["encuesta"], errors="coerce")
    df["periodo"] = pd.to_datetime(df["periodo"], errors="coerce")
    df = df.dropna(subset=["encuesta", "periodo"])
    df["mes_encuesta"] = df["encuesta"].dt.to_period("M")
    df["mes_periodo"] = df["periodo"].dt.to_period("M")
    df = df[df["mes_periodo"] == df["mes_encuesta"] + 1].copy()
    df["fecha"] = df["mes_periodo"].dt.to_timestamp("M")
    df["valor"] = pd.to_numeric(df["mediana"], errors="coerce")
    df["variable"] = df["variable"].astype(str).str.strip()
    df["referencia"] = df["referencia"].astype(str).str.strip()
    df = df.dropna(subset=["valor"]).sort_values("fecha").drop_duplicates(subset=["variable", "referencia", "fecha"])
    return df[["variable", "referencia", "fecha", "valor"]].reset_index(drop=True)


def fetch_rem_variable(variable: str, referencia: str, start_date: str | None = None) -> pd.DataFrame:
    """
    Serie de "expectativa a un mes vista, siempre la más reciente disponible
    para cada mes" para la variable/referencia pedida (tal como figuran en
    las columnas "Variable"/"Referencia" de la hoja, ej. "Precios minoristas
    (IPC nivel general; INDEC)" / "var. % mensual"). El resultado queda
    indexado por Período (el mes al que corresponde la expectativa, no la
    fecha de la encuesta), normalizado a fin de mes para alinear con
    "Inflación mensual (IPC)" (ArgentinaDatos) en los gráficos de overlay.

    El Excel del BCRA se actualiza una vez por mes (una encuesta nueva) pero
    pesa ~1,5 MB; para no descargarlo de nuevo en cada corrida diaria del
    pipeline sin necesidad (tráfico contra el servidor del BCRA sin ningún
    dato nuevo la gran mayoría de los días), la tabla ya extraída (todas las
    variables, no sólo ésta) se cachea en data/_cache_rem.csv con una marca
    de fecha en data/_cache_rem_meta.json: sólo se vuelve a descargar si la
    caché tiene más de REM_FRESCURA_DIAS días o todavía no tiene la variable
    pedida (ej. la primera vez que se agrega un indicador REM nuevo).
    """
    completo = None
    if _REM_CACHE_META.exists() and _REM_CACHE_CSV.exists():
        try:
            meta = json.loads(_REM_CACHE_META.read_text(encoding="utf-8"))
            descargado = pd.to_datetime(meta.get("descargado"))
            if (pd.Timestamp.today().normalize() - descargado).days < REM_FRESCURA_DIAS:
                cache = pd.read_csv(_REM_CACHE_CSV, parse_dates=["fecha"])
                if ((cache["variable"] == variable) & (cache["referencia"] == referencia)).any():
                    completo = cache
        except (ValueError, KeyError, OSError, json.JSONDecodeError):
            completo = None

    if completo is None:
        completo = _descargar_y_transformar_rem()
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        completo.to_csv(_REM_CACHE_CSV, index=False)
        _REM_CACHE_META.write_text(
            json.dumps({"descargado": pd.Timestamp.today().normalize().isoformat()}), encoding="utf-8")

    sub = completo[(completo["variable"] == variable) & (completo["referencia"] == referencia)]
    if sub.empty:
        return pd.DataFrame(columns=["fecha", "valor"])
    df = sub[["fecha", "valor"]].sort_values("fecha").drop_duplicates(subset="fecha").reset_index(drop=True)
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 6) BCRA: componentes de "reservas ajustadas" (swap China + organismos internacionales)
# ---------------------------------------------------------------------------
BCRA_BALANCE_XLS_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/Serieanual.xls"
BCRA_SDDS_URL_TPL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/temp{mes:02d}{anio2:02d}.pdf"
ORGANISMOS_FRESCURA_DIAS = 7  # el balance semanal se actualiza ~4 veces por mes, no hace falta bajarlo a diario
SWAP_CHINA_INICIO = pd.Timestamp("2022-12-31")  # antes de esta fecha el swap no tiene fila propia en la planilla SDDS
_ORGANISMOS_CACHE_CSV = _CACHE_DIR / "_cache_organismos_internacionales.csv"
_ORGANISMOS_CACHE_META = _CACHE_DIR / "_cache_organismos_internacionales_meta.json"
_SWAP_CACHE_CSV = _CACHE_DIR / "_cache_swap_china.csv"


def fetch_bcra_organismos_internacionales() -> pd.DataFrame:
    """
    "Obligaciones con organismos internacionales" (BCRA, dataset del Balance
    Semanal -- Serie Anual de Balances Semanales, un único Excel histórico
    desde 1998, se actualiza ~4 veces por mes). El propio BCRA documenta este
    rubro como "las operaciones y cuentas de depósito del F.M.I., Banco
    Internacional de Pagos de Basilea (B.I.S.) y otros organismos" (más el
    Uso del Tramo de Reservas y su contrapartida) -- es un AGREGADO FMI+BIS+
    otros, no BIS aislado. Viene en miles de $ y se convierte a millones de
    USD con el tipo de cambio de referencia de la misma planilla.

    El archivo pesa ~1,8 MB: se cachea (mismo mecanismo que el REM) para no
    volver a descargarlo en cada corrida diaria sin necesidad.
    """
    completo = None
    if _ORGANISMOS_CACHE_META.exists() and _ORGANISMOS_CACHE_CSV.exists():
        try:
            meta = json.loads(_ORGANISMOS_CACHE_META.read_text(encoding="utf-8"))
            descargado = pd.to_datetime(meta.get("descargado"))
            if (pd.Timestamp.today().normalize() - descargado).days < ORGANISMOS_FRESCURA_DIAS:
                completo = pd.read_csv(_ORGANISMOS_CACHE_CSV, parse_dates=["fecha"])
        except (ValueError, KeyError, OSError, json.JSONDecodeError):
            completo = None

    if completo is None:
        completo = _descargar_y_transformar_organismos()
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        completo.to_csv(_ORGANISMOS_CACHE_CSV, index=False)
        _ORGANISMOS_CACHE_META.write_text(
            json.dumps({"descargado": pd.Timestamp.today().normalize().isoformat()}), encoding="utf-8")
    return completo[["fecha", "valor"]].reset_index(drop=True)


def _descargar_y_transformar_organismos() -> pd.DataFrame:
    r = _get(BCRA_BALANCE_XLS_URL)
    wb = xlrd.open_workbook(file_contents=r.content)
    filas = []
    for nombre_hoja in wb.sheet_names():
        if "serie semanal" not in nombre_hoja.lower():
            continue
        ws = wb.sheet_by_name(nombre_hoja)
        # Las filas se buscan por ETIQUETA (no por índice fijo): la posición cambia de año a año
        # dentro del mismo archivo. "startswith" (no "in") porque el activo tiene un rubro parecido
        # ("- Pago Obligaciones con Organismos Internacionales", un adelanto al Gobierno Nacional,
        # NADA que ver) que un simple "contiene" matchearía por error.
        fila_organismos = fila_tc = fila_fechas = None
        for r_ in range(ws.nrows):
            etiqueta = str(ws.cell_value(r_, 0)).strip().upper()
            if etiqueta.startswith("OBLIGACIONES CON ORGANISMOS") and fila_organismos is None:
                fila_organismos = r_
            if etiqueta == "TIPO DE CAMBIO" and fila_tc is None:
                fila_tc = r_
        for r_ in range(ws.nrows):
            v = ws.cell_value(r_, 1)
            if isinstance(v, float) and 20000 < v < 60000:  # rango plausible de fecha serial de Excel
                fila_fechas = r_
                break
        if fila_organismos is None or fila_tc is None or fila_fechas is None:
            continue
        for c in range(1, ws.ncols):
            serial, miles_pesos, tc = (ws.cell_value(fila_fechas, c), ws.cell_value(fila_organismos, c),
                                        ws.cell_value(fila_tc, c))
            if not (isinstance(serial, float) and isinstance(miles_pesos, float)
                    and isinstance(tc, float) and tc > 0):
                continue
            fecha = xlrd.xldate.xldate_as_datetime(serial, wb.datemode)
            filas.append({"fecha": fecha, "valor": miles_pesos * 1000 / tc / 1e6})
    df = pd.DataFrame(filas, columns=["fecha", "valor"])
    return df.drop_duplicates(subset="fecha").sort_values("fecha").reset_index(drop=True)


def _parse_num_ar(s) -> float | None:
    """'-17.869,31' -> -17869.31 (formato numérico argentino: punto de miles, coma decimal)."""
    if not s or not str(s).strip():
        return None
    try:
        return float(str(s).strip().replace(".", "").replace(",", "."))
    except ValueError:
        return None


def _leer_swap_china_mes(periodo: pd.Period) -> float | None:
    """
    Descarga (si existe) la planilla mensual "Reservas Internacionales/Liquidez
    en Moneda Extranjera" (formato estándar SDDS del FMI) de un mes puntual y
    busca la fila "swaps de monedas" (sección II.2) en cualquier página de la
    tabla -- devuelve None si el mes todavía no se publicó (404) o si el PDF
    no tiene esa fila (versiones viejas, antes de dic-2022, donde el swap
    estaba mezclado con otra sección).
    """
    url = BCRA_SDDS_URL_TPL.format(mes=periodo.month, anio2=periodo.year % 100)
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if r.status_code != 200 or not r.content.startswith(b"%PDF"):
            return None
    except requests.RequestException:
        return None
    try:
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            for page in pdf.pages:
                for tabla in page.extract_tables():
                    for fila in tabla:
                        etiqueta = (fila[0] or "").replace("\n", " ").lower()
                        if "swaps de monedas" in etiqueta:
                            valor = _parse_num_ar(fila[2] if len(fila) > 2 else None)
                            if valor is not None:
                                return abs(valor)
    except Exception:
        return None
    return None


def fetch_bcra_swap_china() -> pd.DataFrame:
    """
    Posición de swap de monedas BCRA-PBOC (Banco Popular de China), sección
    II.2 ("swaps de monedas") de la planilla mensual SDDS del BCRA. Sólo
    tiene fila propia desde el cierre de dic-2022 (antes estaba mezclado con
    pases en otra sección, según nota del propio BCRA); no se rellena hacia
    atrás con ninguna otra fuente -- la serie arranca ahí y punto.

    Cachea cada mes ya conseguido de forma PERMANENTE en un CSV (los meses
    ya publicados no cambian retroactivamente): en cada corrida sólo intenta
    bajar los meses que todavía faltan en la caché (normalmente el último,
    a veces ninguno si el BCRA no publicó todavía el mes en curso).
    """
    cache = pd.DataFrame(columns=["fecha", "valor"])
    if _SWAP_CACHE_CSV.exists():
        cache = pd.read_csv(_SWAP_CACHE_CSV, parse_dates=["fecha"])

    meses_tenidos = set(cache["fecha"].dt.to_period("M")) if not cache.empty else set()
    mes_cursor = SWAP_CHINA_INICIO.to_period("M")
    mes_final = pd.Timestamp.today().to_period("M")
    nuevas = []
    while mes_cursor <= mes_final:
        if mes_cursor not in meses_tenidos:
            valor = _leer_swap_china_mes(mes_cursor)
            if valor is not None:
                nuevas.append({"fecha": mes_cursor.to_timestamp("M"), "valor": valor})
        mes_cursor += 1

    if nuevas:
        cache = pd.DataFrame(nuevas) if cache.empty else pd.concat([cache, pd.DataFrame(nuevas)], ignore_index=True)
        cache = cache.drop_duplicates(subset="fecha").sort_values("fecha")
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache.to_csv(_SWAP_CACHE_CSV, index=False)
    return cache[["fecha", "valor"]].sort_values("fecha").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 7) BCRA: morosidad (cartera irregular) de Familias por línea de crédito
# ---------------------------------------------------------------------------
INFBANC_ANEXO_URL = "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/informes/InfBanc_Anexo.xlsx"
MOROSIDAD_LINEAS_FRESCURA_DIAS = 7  # el Informe sobre Bancos se actualiza ~1 vez por mes
_MOROSIDAD_LINEAS_CACHE_CSV = _CACHE_DIR / "_cache_morosidad_lineas.csv"
_MOROSIDAD_LINEAS_CACHE_META = _CACHE_DIR / "_cache_morosidad_lineas_meta.json"
MOROSIDAD_LINEAS_SERIES = ("Cartera irregular total", "Personales", "Tarjetas de crédito")


def _sin_acentos(s: str) -> str:
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _descargar_y_transformar_morosidad_lineas() -> pd.DataFrame:
    """
    Anexo estadístico (Excel) del "Informe sobre Bancos" mensual del BCRA, hoja
    "Calidad de Cartera (por líneas)": tiene una sección fija "2. Familias -
    Total" con el ratio de irregularidad (%) de la cartera de Familias, y sus
    líneas Personales y Tarjetas de crédito por separado -- exactamente el
    desagregado que el BCRA NO publica en datos.gob.ar (ahí sólo está por tipo
    de banco, no por tipo de deudor/línea).

    Las filas se buscan por ETIQUETA (no por índice fijo, puede moverse de
    versión a versión del archivo), acotadas a la sección "2. Familias -
    Total" (entre esa etiqueta y la siguiente sección "2.1. Familias - En
    UVA") para no confundirlas con las mismas etiquetas que también aparecen
    en la sección "1. Total Sector Privado" (Familias + Empresas) más arriba.
    """
    r = _get(INFBANC_ANEXO_URL)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    hoja = next(s for s in wb.sheetnames if _sin_acentos(s.lower()).startswith("calidad de cartera (por l"))
    ws = wb[hoja]

    inicio = fin = None
    for row in range(1, ws.max_row + 1):
        etiqueta = _sin_acentos(str(ws.cell(row, 1).value or "").strip().lower())
        if etiqueta.startswith("2. familias") and inicio is None:
            inicio = row
        elif etiqueta.startswith("2.1") and inicio is not None:
            fin = row
            break
    if inicio is None:
        return pd.DataFrame(columns=["fecha", "serie", "valor"])
    fin = fin or ws.max_row

    fila_fechas = fin_tabla = None
    for row in range(inicio, fin):
        etiqueta = _sin_acentos(str(ws.cell(row, 1).value or "").strip().lower())
        if etiqueta == "en porcentaje" and fila_fechas is None:
            fila_fechas = row
        elif etiqueta.startswith("fuente:") and fila_fechas is not None:
            # "Fuente: BCRA" cierra la sub-tabla "Ratio de irregularidad" -- sin este corte, el
            # scan seguiría de largo hasta la sección "2.1" y también agarraría la sub-tabla
            # "Saldo total de financiaciones" más abajo, que reusa las MISMAS etiquetas de fila
            # ("Personales", "Tarjetas de crédito") pero en millones de $, no en % -- se mezclaban
            # ambas bajo la misma serie y quedaban valores de miles de millones donde debía haber
            # un ratio de 0-100.
            fin_tabla = row
            break
    if fila_fechas is None:
        return pd.DataFrame(columns=["fecha", "serie", "valor"])
    fin_tabla = fin_tabla or fin
    fechas = [ws.cell(fila_fechas, c).value for c in range(2, ws.max_column + 1)]

    filas = []
    for row in range(fila_fechas + 1, fin_tabla):
        etiqueta_cruda = str(ws.cell(row, 1).value or "").strip()
        etiqueta_norm = _sin_acentos(etiqueta_cruda.lower())
        objetivo = next((s for s in MOROSIDAD_LINEAS_SERIES if _sin_acentos(s.lower()) == etiqueta_norm), None)
        if objetivo is None:
            continue
        for c, fecha in enumerate(fechas, start=2):
            if not isinstance(fecha, datetime.datetime):
                continue
            valor = ws.cell(row, c).value
            if isinstance(valor, (int, float)):
                filas.append({"fecha": pd.Timestamp(fecha), "serie": objetivo, "valor": float(valor)})
    return pd.DataFrame(filas)


def fetch_bcra_morosidad_lineas(serie: str, start_date: str | None = None) -> pd.DataFrame:
    """serie: uno de MOROSIDAD_LINEAS_SERIES ('Cartera irregular total', 'Personales', 'Tarjetas de crédito')."""
    completo = None
    if _MOROSIDAD_LINEAS_CACHE_META.exists() and _MOROSIDAD_LINEAS_CACHE_CSV.exists():
        try:
            meta = json.loads(_MOROSIDAD_LINEAS_CACHE_META.read_text(encoding="utf-8"))
            descargado = pd.to_datetime(meta.get("descargado"))
            if (pd.Timestamp.today().normalize() - descargado).days < MOROSIDAD_LINEAS_FRESCURA_DIAS:
                completo = pd.read_csv(_MOROSIDAD_LINEAS_CACHE_CSV, parse_dates=["fecha"])
        except (ValueError, KeyError, OSError, json.JSONDecodeError):
            completo = None

    if completo is None:
        completo = _descargar_y_transformar_morosidad_lineas()
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        completo.to_csv(_MOROSIDAD_LINEAS_CACHE_CSV, index=False)
        _MOROSIDAD_LINEAS_CACHE_META.write_text(
            json.dumps({"descargado": pd.Timestamp.today().normalize().isoformat()}), encoding="utf-8")

    sub = completo[completo["serie"] == serie]
    if sub.empty:
        return pd.DataFrame(columns=["fecha", "valor"])
    df = sub[["fecha", "valor"]].sort_values("fecha").drop_duplicates(subset="fecha").reset_index(drop=True)
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 8) Secretaría de Finanzas: deuda bruta de la Administración Central
# ---------------------------------------------------------------------------
DEUDA_PAGINA_URL = "https://www.argentina.gob.ar/economia/finanzas/datos-mensuales"
DEUDA_FRESCURA_DIAS = 7  # el boletín se actualiza ~1 vez por mes
_DEUDA_CACHE_CSV = _CACHE_DIR / "_cache_deuda_bruta.csv"
_DEUDA_CACHE_META = _CACHE_DIR / "_cache_deuda_bruta_meta.json"


def _descargar_deuda_bruta() -> pd.DataFrame:
    """
    Boletín mensual de deuda bruta de la Administración Central (Secretaría
    de Finanzas, Ministerio de Economía), hoja "A.1", fila "A- DEUDA BRUTA
    ( I + II + III)" (total, millones de USD corrientes), serie mensual
    desde 2019. El nombre del archivo Excel cambia cada mes (ej.
    boletin_mensual_31_07_2026_0.xlsx, con el día de cierre del mes y un
    sufijo que puede variar si hay una revisión posterior) -- no es un
    patrón de URL predecible como el de otros fetchers de este archivo, así
    que se scrapea la página de descarga para encontrar el link vigente en
    vez de intentar adivinar el nombre. La fila de fechas (encabezado) y la
    fila del total se ubican buscando su contenido (no un número de fila
    fijo), para no romperse si el boletín agrega o saca alguna fila arriba.
    """
    pagina = _get(DEUDA_PAGINA_URL)
    m = re.search(r'href="(/sites/default/files/boletin_mensual[^"]*\.xlsx)"', pagina.text)
    if not m:
        raise ValueError("No se encontró el link al boletín mensual de deuda bruta en la página de Hacienda")
    r = _get("https://www.argentina.gob.ar" + m.group(1))
    with io.BytesIO(r.content) as buf:
        df = pd.read_excel(buf, sheet_name="A.1", header=None, engine="openpyxl")

    fila_header = None
    for i in range(min(20, len(df))):
        if sum(isinstance(v, (pd.Timestamp, datetime.datetime)) for v in df.iloc[i]) > 5:
            fila_header = i
            break
    if fila_header is None:
        raise ValueError("No se encontró la fila de fechas en la hoja A.1 del boletín de deuda")
    fechas = df.iloc[fila_header]

    fila_total = None
    for i in range(fila_header, len(df)):
        if any(isinstance(v, str) and v.strip().upper().startswith("A- DEUDA BRUTA") for v in df.iloc[i]):
            fila_total = i
            break
    if fila_total is None:
        raise ValueError("No se encontró la fila 'A- DEUDA BRUTA' en la hoja A.1 del boletín de deuda")
    valores = df.iloc[fila_total]

    # La etiqueta de la fila vive en una columna propia (no siempre la 0): en vez de asumir un
    # offset fijo entre columna de etiqueta y columna del primer dato, se empareja cada columna
    # de 'fechas' con la misma columna de 'valores' y sólo se queda con los pares donde AMBAS
    # celdas tienen el tipo esperado (fecha real / número real) -- así se saltean solas las
    # columnas de etiqueta (texto o vacías) sin necesidad de saber en qué columna están.
    filas = []
    for col in range(len(fechas)):
        f, v = fechas.iloc[col], valores.iloc[col]
        if isinstance(f, (pd.Timestamp, datetime.datetime)) and isinstance(v, (int, float)) and pd.notna(v):
            filas.append({"fecha": pd.Timestamp(f), "valor": float(v)})
    return pd.DataFrame(filas).sort_values("fecha").reset_index(drop=True)


def fetch_deuda_bruta(start_date: str | None = None) -> pd.DataFrame:
    """Deuda bruta de la Administración Central, cacheada igual que REM/organismos
    internacionales/morosidad por líneas (máximo 1 descarga por semana)."""
    cache = None
    if _DEUDA_CACHE_META.exists() and _DEUDA_CACHE_CSV.exists():
        try:
            meta = json.loads(_DEUDA_CACHE_META.read_text(encoding="utf-8"))
            descargado = pd.to_datetime(meta.get("descargado"))
            if (pd.Timestamp.today().normalize() - descargado).days < DEUDA_FRESCURA_DIAS:
                cache = pd.read_csv(_DEUDA_CACHE_CSV, parse_dates=["fecha"])
        except (ValueError, KeyError, OSError, json.JSONDecodeError):
            cache = None

    if cache is None:
        cache = _descargar_deuda_bruta()
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache.to_csv(_DEUDA_CACHE_CSV, index=False)
        _DEUDA_CACHE_META.write_text(
            json.dumps({"descargado": pd.Timestamp.today().normalize().isoformat()}), encoding="utf-8")

    df = cache[["fecha", "valor"]].sort_values("fecha").drop_duplicates(subset="fecha").reset_index(drop=True)
    if start_date:
        df = df[df["fecha"] >= pd.to_datetime(start_date)]
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 5) BYMA -> índices históricos (MERVAL, BURCAP)
# ---------------------------------------------------------------------------
_BYMA_BASE = "https://open.bymadata.com.ar/vanoms-be-core/rest/api/bymadata/free"


def fetch_byma_indice(symbol: str, start_date: str | None = None) -> pd.DataFrame:
    """
    Serie histórica de un índice de BYMA (symbol="M" -> S&P MERVAL, "G" -> BURCAP).

    El certificado de BYMA tiene una CA intermedia que no está en el bundle
    estándar de certifi -- se pide con verify=False, mismo criterio que las
    librerías públicas de BYMA (pyhomebroker, bymadata). La conexión sigue
    siendo TLS, sólo se salta la validación de cadena.
    """
    desde_dt = pd.to_datetime(start_date) if start_date else pd.Timestamp("2024-01-01")
    desde_unix = int(desde_dt.timestamp())
    hasta_unix = int(pd.Timestamp.today().timestamp())
    url = f"{_BYMA_BASE}/chart/index-historical-series/history"
    params = {"symbol": symbol, "resolution": "D", "from": desde_unix, "to": hasta_unix}

    ultimo_error = None
    for intento in range(3):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=TIMEOUT, verify=False)
            r.raise_for_status()
            data = r.json()
            break
        except requests.RequestException as e:
            ultimo_error = e
            time.sleep(2 * (intento + 1))
    else:
        print(f"  [ADVERTENCIA] fetch_byma_indice symbol={symbol}: {ultimo_error}")
        return pd.DataFrame(columns=["fecha", "valor"])

    ts = data.get("t") or []
    cierres = data.get("c") or []
    if not ts or not cierres:
        return pd.DataFrame(columns=["fecha", "valor"])
    df = pd.DataFrame({
        "fecha": pd.to_datetime(ts, unit="s").normalize(),
        "valor": [float(v) for v in cierres],
    })
    return df.sort_values("fecha").drop_duplicates(subset="fecha").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 6) MAE -> curva de dólar futuro (DDF), snapshot del día (no serie fecha/valor)
# ---------------------------------------------------------------------------
_MAE_BASE = "https://api.marketdata.mae.com.ar/api"


def fetch_mae_ddf() -> list[dict]:
    """
    Curva completa de dólar futuro (DDF) del día: un contrato por vencimiento
    mensual. Endpoint no documentado oficialmente por el MAE ("puede cambiar
    sin aviso", datos delayed 5-15 min intradía) -- no es una serie fecha/valor
    como el resto de los fetchers, es una foto de todos los contratos vigentes
    en la fecha de la corrida, así que no pasa por traer()/el histórico
    fecha/valor: se maneja aparte en la pestaña financiera.
    """
    try:
        r = _get(f"{_MAE_BASE}/mercado/resumen/DDF")
        data = r.json()
    except (RuntimeError, ValueError) as e:
        print(f"  [ADVERTENCIA] fetch_mae_ddf: {e}")
        return []
    contratos = data if isinstance(data, list) else data.get("data", [])
    filas = []
    for c in contratos:
        try:
            filas.append({
                "ticker": c["ticker"],
                "ultimo": float(c["ultimo"]),
                "variacion": float(c.get("variacion") or 0),
            })
        except (KeyError, ValueError, TypeError):
            pass
    return filas


# ---------------------------------------------------------------------------
# Dispatcher: elige el fetcher según el config del indicador
# ---------------------------------------------------------------------------
def traer(indicador: dict, start_date: str | None = None) -> pd.DataFrame:
    fuente = indicador["fuente"]
    if fuente == "datos_gob":
        return fetch_datos_gob(indicador["id"], start_date)
    if fuente == "argentinadatos":
        return fetch_argentinadatos(indicador["endpoint"], start_date)
    if fuente == "dolar":
        return fetch_dolar(indicador["casa"], start_date)
    if fuente == "bcra":
        return fetch_bcra(indicador["id_variable"], start_date)
    if fuente == "rem_bcra":
        return fetch_rem_variable(indicador["variable"], indicador["referencia"], start_date)
    if fuente == "bcra_morosidad_lineas":
        return fetch_bcra_morosidad_lineas(indicador["serie"], start_date)
    if fuente == "deuda_bruta":
        return fetch_deuda_bruta(start_date)
    if fuente == "byma":
        return fetch_byma_indice(indicador["symbol"], start_date)
    raise ValueError(f"Fuente desconocida: {fuente}")
